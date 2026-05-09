import csv
import io
from typing import Any

TITLE_ALIASES = {"课程名", "课程", "名称", "title", "course", "name"}
TEACHER_ALIASES = {"教师", "老师", "teacher", "instructor"}
EXAM_DATE_ALIASES = {"考试日期", "考试时间", "exam_date", "exam date", "examdate"}


def _find_column(headers: list[str], aliases: set[str]) -> str | None:
    for h in headers:
        if h.strip().lower() in aliases:
            return h
    return None


def parse_csv(content: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(content))
    headers = list(reader.fieldnames or [])
    title_col = _find_column(headers, TITLE_ALIASES)
    if title_col is None:
        raise ValueError("CSV 缺少课程名列（课程名/课程/名称/title）")
    teacher_col = _find_column(headers, TEACHER_ALIASES)
    exam_date_col = _find_column(headers, EXAM_DATE_ALIASES)

    rows: list[dict] = []
    for row in reader:
        title = row.get(title_col, "").strip()
        if not title:
            continue
        teacher = row[teacher_col].strip() if teacher_col and row.get(teacher_col) else None
        exam_date = row[exam_date_col].strip() if exam_date_col and row.get(exam_date_col) else None
        rows.append({"title": title, "teacher": teacher, "exam_date": exam_date})
    return rows


def parse_excel(content: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 文件没有活动工作表")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]
    title_col = _find_column(headers, TITLE_ALIASES)
    if title_col is None:
        wb.close()
        raise ValueError("Excel 缺少课程名列（课程名/课程/名称/title）")
    title_idx = headers.index(title_col)

    teacher_col = _find_column(headers, TEACHER_ALIASES)
    teacher_idx = headers.index(teacher_col) if teacher_col else None

    exam_date_col = _find_column(headers, EXAM_DATE_ALIASES)
    exam_date_idx = headers.index(exam_date_col) if exam_date_col else None

    results: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None:
            continue
        title = str(row[title_idx] or "").strip() if title_idx < len(row) else ""
        if not title or title == "None":
            continue
        teacher = (
            str(row[teacher_idx] or "").strip()
            if teacher_idx is not None and teacher_idx < len(row)
            else None
        )
        exam_date = (
            str(row[exam_date_idx] or "").strip()
            if exam_date_idx is not None and exam_date_idx < len(row)
            else None
        )
        if teacher in ("None", "none", ""):
            teacher = None
        if exam_date in ("None", "none", ""):
            exam_date = None
        results.append({"title": title, "teacher": teacher, "exam_date": exam_date})
    wb.close()
    return results
